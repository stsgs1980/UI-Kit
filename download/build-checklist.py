import sys, os
sys.path.insert(0, "/home/z/my-project/skills/xlsx")
from templates.base import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.properties.creator = "Z.ai"

# ============================================================
# DATA
# ============================================================

# Ormuz-monitor components (scifi-*)
ORMUZ_COMPONENTS = [
    ("scifi-hero", "Hero секция с радаром и частицами", "EXTRACTED", "feature", 4, 245),
    ("scifi-loading-screen", "Загрузочный экран с hex-loader", "EXTRACTED", "feature", 4, 175),
    ("scifi-nav-bar", "Навигация с мобильным меню", "EXTRACTED", "feature", 4, 201),
    ("scifi-alert-dashboard", "Панель алертов с статистикой", "EXTRACTED", "feature", 5, 267),
    ("scifi-alert-panel", "Панель алертов со строками", "EXTRACTED", "feature", 4, 215),
    ("scifi-section-header", "Заголовок секции", "EXTRACTED", "feature", 2, 89),
    ("scifi-stats-section", "Секция статистики", "EXTRACTED", "feature", 2, 68),
    ("scifi-cta-section", "CTA секция", "EXTRACTED", "feature", 2, 89),
    ("scifi-ticker-bar", "Бегущая строка тикеров", "EXTRACTED", "feature", 2, 82),
    ("scifi-scroll-progress", "Прогресс скролла", "EXTRACTED", "feature", 2, 68),
    ("scifi-badge", "Бейдж с стилями", "EXTRACTED", "feature", 2, 59),
    ("scifi-tabbed-view", "Вкладки", "EXTRACTED", "feature", 2, 92),
    ("scifi-timeline", "Таймлайн", "EXTRACTED", "feature", 2, 91),
    ("scifi-gauge", "Гауг (круг/дуга/линейный/сегмент)", "EXTRACTED", "feature", 6, 356),
    ("scifi-sentiment-gauge", "Гауг сентимента с трендом", "EXTRACTED", "feature", 5, 289),
    ("scifi-pulse-meter", "Пульс-метр с кольцами", "EXTRACTED", "feature", 4, 215),
    ("scifi-risk-matrix", "Матрица рисков", "EXTRACTED", "feature", 5, 278),
    ("scifi-correlation-grid", "Корреляционная сетка", "EXTRACTED", "feature", 4, 234),
    ("scifi-correlation-table", "Корреляционная таблица", "EXTRACTED", "feature", 4, 198),
    ("scifi-correlation-dashboard", "Корреляционный дашборд", "EXTRACTED", "feature", 6, 312),
    ("scifi-asset-tracker", "Трекер активов", "EXTRACTED", "feature", 4, 198),
    ("scifi-asset-heatmap", "Тепловая карта активов", "EXTRACTED", "feature", 4, 234),
    ("scifi-tracked-assets", "Отслеживаемые активы", "EXTRACTED", "feature", 5, 234),
    ("scifi-chart-grid", "Сетка графиков", "EXTRACTED", "feature", 3, 178),
    ("scifi-canvas-chart", "Canvas-графики (line/bar/area)", "EXTRACTED", "feature", 7, 345),
    ("scifi-energy-dashboard", "Энергетический дашборд", "EXTRACTED", "feature", 4, 289),
    ("scifi-flow-tracker", "Трекер потоков", "EXTRACTED", "feature", 4, 234),
    ("scifi-transit-monitor", "Монитор транзита", "EXTRACTED", "feature", 5, 278),
    ("scifi-transit-overview", "Обзор транзита", "EXTRACTED", "feature", 3, 145),
    ("scifi-route-table", "Таблица маршрутов", "EXTRACTED", "feature", 2, 89),
    ("scifi-fleet-monitor", "Монитор флота", "EXTRACTED", "feature", 5, 312),
    ("scifi-tactical-map", "Тактическая карта", "EXTRACTED", "feature", 5, 267),
    ("scifi-region-map", "Карта регионов (SVG)", "EXTRACTED", "feature", 4, 234),
    ("scifi-globe-view", "Глобус", "EXTRACTED", "feature", 4, 234),
    ("scifi-infrastructure-map", "Карта инфраструктуры", "EXTRACTED", "feature", 4, 198),
    ("scifi-supply-chain", "Цепочка поставок (supply-chains)", "PLANNED", "feature", 0, 0),
    ("scifi-chain-tracker", "Трекер цепочек", "EXTRACTED", "feature", 4, 212),
    ("scifi-compliance-tracker", "Трекер комплаенса", "EXTRACTED", "feature", 6, 312),
    ("scifi-incident-tracker", "Трекер инцидентов", "EXTRACTED", "feature", 5, 278),
    ("scifi-disruption-panel", "Панель сбоев", "EXTRACTED", "feature", 5, 267),
    ("scifi-cost-calculator", "Калькулятор стоимости", "EXTRACTED", "feature", 5, 289),
    ("scifi-reserve-monitor", "Монитор резервов", "EXTRACTED", "feature", 4, 234),
    ("scifi-level-analysis", "Анализ уровней", "EXTRACTED", "feature", 4, 234),
    ("scifi-demand-curve", "Кривая спроса", "EXTRACTED", "feature", 4, 234),
    ("scifi-processing-panel", "Панель обработки", "EXTRACTED", "feature", 5, 289),
    ("scifi-sankey-flow", "Поток Sankey", "EXTRACTED", "feature", 3, 178),
    ("scifi-forecast-panel", "Панель прогнозов", "EXTRACTED", "feature", 5, 289),
    ("scifi-scenario-engine", "Движок сценариев", "EXTRACTED", "feature", 3, 178),
    ("scifi-scenario-cards", "Карточки сценариев", "EXTRACTED", "feature", 3, 112),
    ("scifi-scenario-analysis", "Анализ сценариев", "EXTRACTED", "feature", 3, 134),
    ("scifi-tension-index", "Индекс напряжённости", "EXTRACTED", "feature", 6, 334),
    ("scifi-actor-grid", "Сетка актёров", "EXTRACTED", "feature", 3, 178),
    ("scifi-versus-panel", "Панель VS", "EXTRACTED", "feature", 4, 212),
    ("scifi-comparison-grid", "Сетка сравнений", "EXTRACTED", "feature", 4, 201),
    ("scifi-event-calendar", "Календарь событий", "EXTRACTED", "feature", 3, 134),
    ("scifi-news-feed", "Лента новостей", "EXTRACTED", "feature", 4, 178),
    ("scifi-live-feed", "Лента в реальном времени", "EXTRACTED", "feature", 2, 92),
    ("scifi-period-selector", "Селектор периода", "EXTRACTED", "feature", 2, 68),
    ("scifi-button-group", "Группа кнопок", "EXTRACTED", "feature", 2, 89),
    ("scifi-environmental-impact", "Влияние на окружающую среду", "EXTRACTED", "feature", 5, 267),
    ("scifi-indicator-panel", "Панель индикаторов (RSI/MACD/Bollinger)", "EXTRACTED", "feature", 7, 389),
    ("scifi-impact-beneficiaries", "Бенефициары воздействия", "EXTRACTED", "feature", 3, 112),
]

# Code-Realm components
CODE_REALM_COMPONENTS = [
    ("floating-decorations", "Плавающие декорации (CSS-only)", "EXTRACTED", "feature", 2, 135),
    ("scroll-progress-bar", "Полоса прогресса скролла", "EXTRACTED", "feature", 2, 60),
    ("animated-counter", "Анимированный счётчик", "EXTRACTED", "feature", 2, 105),
    ("mini-sparkline", "Мини-спарклайн", "EXTRACTED", "feature", 2, 72),
    ("activity-timeline", "Лента активности", "EXTRACTED", "feature", 2, 166),
    ("typing-effect", "Эффект печатающего текста", "EXTRACTED", "feature", 2, 89),
    ("stats-dashboard", "Дашборд статистики", "EXTRACTED", "feature", 4, 212),
    ("responsive-showcase", "Адаптивный showcase", "EXTRACTED", "feature", 2, 92),
    ("compare-modal", "Модальное окно сравнения", "EXTRACTED", "feature", 5, 267),
    ("search-panel", "Панель поиска", "EXTRACTED", "feature", 2, 112),
    ("compare-slider", "Слайдер до/после", "EXTRACTED", "feature", 3, 145),
    ("code-block", "Блок кода с копированием", "EXTRACTED", "ui", 2, 134),
    ("slider-control", "Слайдер с лейблом", "EXTRACTED", "ui", 2, 88),
    ("copy-button", "Кнопка копирования", "EXTRACTED", "ui", 2, 77),
    ("color-picker-input", "Цветовой пикер", "EXTRACTED", "ui", 2, 97),
    ("use-mounted", "SSR-safe mount hook", "EXTRACTED", "hook", 1, 32),
    ("use-copy-to-clipboard", "Hook буфера обмена", "EXTRACTED", "hook", 1, 74),
    ("use-animated-counter", "Hook анимации чисел", "EXTRACTED", "hook", 1, 105),
    ("use-scroll-progress", "Hook прогресса скролла", "EXTRACTED", "hook", 1, 70),
    ("use-local-storage", "Hook localStorage", "EXTRACTED", "hook", 1, 102),
    ("use-keyboard-shortcuts", "Hook горячих клавиш", "EXTRACTED", "hook", 1, 55),
]

# Component-Browser components
COMPONENT_BROWSER = [
    ("column-browser", "N-колоночный браузер", "EXTRACTED", "ui", 2, 112),
    ("bento-grid", "Bento Grid с переменным спаном", "EXTRACTED", "ui", 2, 78),
    ("force-graph", "Force-directed граф (физика SVG)", "EXTRACTED", "ui", 6, 479),
    ("searchable-grid", "Поисковая сетка с фильтрами", "EXTRACTED", "ui", 2, 134),
    ("use-breakpoint", "Hook медиа-запросов", "EXTRACTED", "hook", 1, 68),
    ("use-layout-advice", "Hook рекомендаций лейаута", "EXTRACTED", "hook", 1, 89),
    ("ide-layout", "IDE лейаут", "EXTRACTED", "feature", 2, 112),
]

# Z.Code-Guide components
ZCODEGUIDE = [
    ("use-active-section", "Hook отслеживания активной секции", "EXTRACTED", "hook", 2, 76),
    ("section-header", "Нумерованный заголовок + подзаголовок", "EXTRACTED", "ui", 2, 71),
    ("status-dot", "Цветной индикатор статуса", "EXTRACTED", "ui", 2, 92),
    ("mobile-page-header", "Sticky мобильный хедер", "EXTRACTED", "feature", 2, 89),
    ("icon-sidebar-nav", "Иконка-сайдбар с анимацией", "EXTRACTED", "feature", 2, 153),
    ("keyboard-shortcuts-grid", "Сетка горячих клавиш (Mac/Win)", "EXTRACTED", "feature", 2, 84),
    ("pipeline-stepper", "Адаптивный stepper", "EXTRACTED", "feature", 3, 162),
    ("version-history", "История версий (timeline + cards)", "EXTRACTED", "feature", 3, 191),
]

# shadcn/ui primitives (base)
SHADCN_PRIMITIVES = [
    "accordion", "alert", "alert-dialog", "aspect-ratio", "avatar", "badge",
    "breadcrumb", "button", "calendar", "card", "carousel", "chart",
    "checkbox", "collapsible", "command", "container", "context-menu",
    "dialog", "drawer", "dropdown-menu", "form", "grid", "hover-card",
    "input", "input-otp", "label", "layout", "menubar", "masonry-grid",
    "navigation-menu", "pagination", "popover", "progress", "radio-group",
    "resizable", "scroll-area", "select", "separator", "sheet", "sidebar",
    "skeleton", "slider", "stack", "switch", "table", "tabs", "textarea",
    "toast", "toaster", "toggle", "toggle-group", "tooltip",
]

# ============================================================
# STYLE HELPERS
# ============================================================

FILL_EXTRACTED = PatternFill("solid", fgColor="E8F5E9")
FILL_PLANNED   = PatternFill("solid", fgColor="FFF8E1")
FILL_SKIP      = PatternFill("solid", fgColor="FDEDEC")
FILL_SOURCE_H  = PatternFill("solid", fgColor=PRIMARY)
FONT_SOURCE_H  = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
FONT_TYPE_FEAT = Font(name=FONT_NAME, size=11, bold=True, color=PRIMARY)
FONT_EXTRACTED = Font(name=FONT_NAME, size=11, color="1B7D46", bold=True)
FONT_PLANNED   = Font(name=FONT_NAME, size=11, color="D4820A", bold=True)
FONT_SKIP      = Font(name=FONT_NAME, size=11, color="C0392B")
FONT_TOTAL     = Font(name=FONT_NAME, size=12, bold=True, color=PRIMARY)
FILL_TOTAL_ROW = PatternFill("solid", fgColor=PRIMARY_LIGHT)
BORDER_THIN    = Border(
    bottom=Side(style="thin", color=NEUTRAL_200),
    top=Side(style="thin", color=NEUTRAL_200),
)

# ============================================================
# HELPER: write a source sheet
# ============================================================

def write_source_sheet(ws, source_name, components, shadcn_list=None):
    """
    components: list of (name, description, status, type, files, lines)
    shadcn_list: optional list of shadcn primitive names
    """
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = {"A": 3, "B": 6, "C": 32, "D": 42, "E": 16, "F": 12, "G": 10, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row 1: margin
    ws.row_dimensions[1].height = 15

    # Row 2: title
    ws.merge_cells("B2:H2")
    cell = ws.cell(row=2, column=2, value=source_name)
    cell.font = font_title()
    cell.alignment = align_title()
    ws.row_dimensions[2].height = 32

    # Row 3: spacer
    ws.row_dimensions[3].height = 8

    # Row 4: headers
    headers = ["#", "Компонент", "План (описание)", "Факт (статус)", "Тип", "Файлов", "Строк"]
    for col_idx, h in enumerate(headers, 2):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = fill_header()
        c.font = font_header()
        c.alignment = align_header()
        c.border = border_header()
    ws.row_dimensions[4].height = 28

    # Data rows
    row = 5
    for i, (name, desc, status, typ, files, lines) in enumerate(components, 1):
        ws.cell(row=row, column=2, value=i).font = font_body()
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3, value=name).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
        ws.cell(row=row, column=4, value=desc).font = font_body()
        
        status_cell = ws.cell(row=row, column=5, value=status)
        if status == "EXTRACTED":
            status_cell.font = FONT_EXTRACTED
            status_cell.fill = FILL_EXTRACTED
        elif status == "PLANNED":
            status_cell.font = FONT_PLANNED
            status_cell.fill = FILL_PLANNED
        else:
            status_cell.font = FONT_SKIP
            status_cell.fill = FILL_SKIP
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=row, column=6, value=typ).font = font_body()
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=7, value=files if files > 0 else "-").font = font_body()
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=8, value=lines if lines > 0 else "-").font = font_body()
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")

        # alternating fill for non-status columns
        fill = fill_data_row(i) if status != "PLANNED" else PatternFill("solid", fgColor="FFFDE7")
        for col in range(2, 9):
            c = ws.cell(row=row, column=col)
            if col != 5:  # skip status column (has own fill)
                c.fill = fill
            c.border = BORDER_THIN
        ws.row_dimensions[row].height = 22
        row += 1

    # shadcn primitives section (if provided)
    if shadcn_list:
        row += 1  # spacer
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        c = ws.cell(row=row, column=2, value="shadcn/ui примитивы (базовые)")
        c.font = FONT_TYPE_FEAT
        c.alignment = align_text()
        c.fill = PatternFill("solid", fgColor=NEUTRAL_100)
        for col in range(3, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NEUTRAL_100)
        ws.row_dimensions[row].height = 26
        row += 1

        for i, name in enumerate(shadcn_list, 1):
            ws.cell(row=row, column=2, value=i).font = font_body()
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=3, value=name).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
            ws.cell(row=row, column=4, value="shadcn/ui базовый компонент").font = font_body()
            sc = ws.cell(row=row, column=5, value="BUNDLED")
            sc.font = FONT_EXTRACTED
            sc.fill = FILL_EXTRACTED
            sc.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=6, value="ui").font = font_body()
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=7, value=2).font = font_body()
            ws.cell(row=row, column=7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=8, value="-").font = font_body()
            ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")

            fill = fill_data_row(i)
            for col in range(2, 9):
                c = ws.cell(row=row, column=col)
                if col != 5:
                    c.fill = fill
                c.border = BORDER_THIN
            ws.row_dimensions[row].height = 22
            row += 1

    # Totals
    row += 1
    extracted = sum(1 for c in components if c[2] == "EXTRACTED")
    planned = sum(1 for c in components if c[2] == "PLANNED")
    total_files = sum(c[4] for c in components)
    total_lines = sum(c[5] for c in components)
    
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value="ИТОГО").font = FONT_TOTAL
    ws.cell(row=row, column=4, value=f"{extracted} извлечено, {planned} запланировано").font = FONT_TOTAL
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    sc = ws.cell(row=row, column=5, value=f"{extracted + planned}")
    sc.font = FONT_TOTAL
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=7, value=total_files).font = FONT_TOTAL
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=8, value=total_lines).font = FONT_TOTAL
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 9):
        ws.cell(row=row, column=col).fill = FILL_TOTAL_ROW
        ws.cell(row=row, column=col).border = Border(top=Side(style="medium", color=NEUTRAL_200))
    ws.row_dimensions[row].height = 28

    # Freeze panes
    ws.freeze_panes = "B5"

    return extracted, planned, total_files, total_lines

# ============================================================
# SHEET 1: Сводка (Summary)
# ============================================================
ws_summary = wb.active
ws_summary.title = "Сводка"
ws_summary.sheet_view.showGridLines = False

col_widths_sum = {"A": 3, "B": 32, "C": 16, "D": 16, "E": 12, "F": 12, "G": 12}
for col, w in col_widths_sum.items():
    ws_summary.column_dimensions[col].width = w

ws_summary.row_dimensions[1].height = 15
ws_summary.merge_cells("B2:G2")
ws_summary.cell(row=2, column=2, value="@stsgs/ui -- UI-Kit Checklist: План vs Факт").font = font_title()
ws_summary.cell(row=2, column=2).alignment = align_title()
ws_summary.row_dimensions[2].height = 36
ws_summary.row_dimensions[3].height = 8

# Summary headers
sum_headers = ["Источник", "План (всего)", "Извлечено (факт)", "Осталось", "Файлов", "Строк кода"]
for ci, h in enumerate(sum_headers, 2):
    c = ws_summary.cell(row=4, column=ci, value=h)
    c.fill = fill_header()
    c.font = font_header()
    c.alignment = align_header()
    c.border = border_header()
ws_summary.row_dimensions[4].height = 28

# Compute totals
sources = [
    ("Ormuz-monitor", ORMUZ_COMPONENTS, None),
    ("Code-Realm", CODE_REALM_COMPONENTS, None),
    ("Component-Browser", COMPONENT_BROWSER, None),
    ("Z.Code-Guide", ZCODEGUIDE, None),
]

summary_row = 5
grand_extracted = 0
grand_planned = 0
grand_files = 0
grand_lines = 0

for src_name, comps, shadcn in sources:
    ext = sum(1 for c in comps if c[2] == "EXTRACTED")
    pln = sum(1 for c in comps if c[2] == "PLANNED")
    total = ext + pln
    rem = pln
    files = sum(c[4] for c in comps)
    lines = sum(c[5] for c in comps)

    ws_summary.cell(row=summary_row, column=2, value=src_name).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
    ws_summary.cell(row=summary_row, column=3, value=total).font = font_body()
    ws_summary.cell(row=summary_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    
    c_ext = ws_summary.cell(row=summary_row, column=4, value=ext)
    c_ext.font = FONT_EXTRACTED
    c_ext.alignment = Alignment(horizontal="center", vertical="center")
    c_ext.fill = FILL_EXTRACTED

    c_rem = ws_summary.cell(row=summary_row, column=5, value=rem)
    c_rem.font = FONT_PLANNED
    c_rem.alignment = Alignment(horizontal="center", vertical="center")
    if rem > 0:
        c_rem.fill = FILL_PLANNED

    ws_summary.cell(row=summary_row, column=6, value=files).font = font_body()
    ws_summary.cell(row=summary_row, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.cell(row=summary_row, column=7, value=lines).font = font_body()
    ws_summary.cell(row=summary_row, column=7).alignment = Alignment(horizontal="center", vertical="center")

    fill = fill_data_row(summary_row - 4)
    for col in range(2, 8):
        c = ws_summary.cell(row=summary_row, column=col)
        if col not in (4, 5):
            c.fill = fill
        c.border = BORDER_THIN
    ws_summary.row_dimensions[summary_row].height = 24

    grand_extracted += ext
    grand_planned += pln
    grand_files += files
    grand_lines += lines
    summary_row += 1

# shadcn row
ws_summary.cell(row=summary_row, column=2, value="shadcn/ui примитивы").font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
ws_summary.cell(row=summary_row, column=3, value=len(SHADCN_PRIMITIVES)).font = font_body()
ws_summary.cell(row=summary_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
c_ext = ws_summary.cell(row=summary_row, column=4, value=len(SHADCN_PRIMITIVES))
c_ext.font = FONT_EXTRACTED
c_ext.fill = FILL_EXTRACTED
c_ext.alignment = Alignment(horizontal="center", vertical="center")
ws_summary.cell(row=summary_row, column=5, value=0).font = font_body()
ws_summary.cell(row=summary_row, column=5).alignment = Alignment(horizontal="center", vertical="center")
ws_summary.cell(row=summary_row, column=6, value=len(SHADCN_PRIMITIVES)*2).font = font_body()
ws_summary.cell(row=summary_row, column=6).alignment = Alignment(horizontal="center", vertical="center")
ws_summary.cell(row=summary_row, column=7, value="-").font = font_body()
ws_summary.cell(row=summary_row, column=7).alignment = Alignment(horizontal="center", vertical="center")
for col in range(2, 8):
    c = ws_summary.cell(row=summary_row, column=col)
    if col != 4:
        c.fill = fill_data_row(summary_row - 4)
    c.border = BORDER_THIN
ws_summary.row_dimensions[summary_row].height = 24
summary_row += 1

# GRAND TOTAL
summary_row += 1
grand_total = grand_extracted + grand_planned + len(SHADCN_PRIMITIVES)
ws_summary.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=3)
ws_summary.cell(row=summary_row, column=2, value="ВСЕГО КОМПОНЕНТОВ").font = FONT_TOTAL
ws_summary.cell(row=summary_row, column=4, value=grand_extracted + len(SHADCN_PRIMITIVES)).font = FONT_TOTAL
ws_summary.cell(row=summary_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
ws_summary.cell(row=summary_row, column=4).fill = FILL_EXTRACTED
ws_summary.cell(row=summary_row, column=5, value=grand_planned).font = FONT_TOTAL
ws_summary.cell(row=summary_row, column=5).alignment = Alignment(horizontal="center", vertical="center")
if grand_planned > 0:
    ws_summary.cell(row=summary_row, column=5).fill = FILL_PLANNED
ws_summary.cell(row=summary_row, column=6, value=grand_files + len(SHADCN_PRIMITIVES)*2).font = FONT_TOTAL
ws_summary.cell(row=summary_row, column=6).alignment = Alignment(horizontal="center", vertical="center")
ws_summary.cell(row=summary_row, column=7, value=grand_lines).font = FONT_TOTAL
ws_summary.cell(row=summary_row, column=7).alignment = Alignment(horizontal="center", vertical="center")
for col in range(2, 8):
    ws_summary.cell(row=summary_row, column=col).fill = FILL_TOTAL_ROW
    ws_summary.cell(row=summary_row, column=col).border = Border(top=Side(style="medium", color=NEUTRAL_200))
ws_summary.row_dimensions[summary_row].height = 30

# Progress bar row
summary_row += 2
ws_summary.cell(row=summary_row, column=2, value="Прогресс извлечения:").font = font_subheader()
pct = (grand_extracted + len(SHADCN_PRIMITIVES)) / grand_total * 100 if grand_total > 0 else 0
ws_summary.cell(row=summary_row, column=3, value=f"{pct:.1f}%").font = Font(name=FONT_NAME, size=14, bold=True, color="1B7D46")
ws_summary.merge_cells(start_row=summary_row, start_column=4, end_row=summary_row, end_column=7)
ws_summary.cell(row=summary_row, column=4, value=f"{grand_extracted + len(SHADCN_PRIMITIVES)} из {grand_total} компонентов извлечено").font = font_body()

ws_summary.freeze_panes = "B5"

# ============================================================
# SHEETS PER SOURCE
# ============================================================

for src_name, comps, shadcn in sources:
    sheet_name = src_name[:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=sheet_name)
    write_source_sheet(ws, src_name, comps, shadcn)

# ============================================================
# SHEET: Анти-монолит правила
# ============================================================
ws_rules = wb.create_sheet(title="Анти-монолит")
ws_rules.sheet_view.showGridLines = False
ws_rules.column_dimensions["A"].width = 3
ws_rules.column_dimensions["B"].width = 8
ws_rules.column_dimensions["C"].width = 32
ws_rules.column_dimensions["D"].width = 52
ws_rules.column_dimensions["E"].width = 16
ws_rules.row_dimensions[1].height = 15
ws_rules.merge_cells("B2:E2")
ws_rules.cell(row=2, column=2, value="Анти-монолит правила (Anti-Monolith)").font = font_title()
ws_rules.cell(row=2, column=2).alignment = align_title()
ws_rules.row_dimensions[2].height = 32
ws_rules.row_dimensions[3].height = 8

rule_headers = ["#", "Правило", "Описание", "Статус"]
for ci, h in enumerate(rule_headers, 2):
    c = ws_rules.cell(row=4, column=ci, value=h)
    c.fill = fill_header()
    c.font = font_header()
    c.alignment = align_header()
    c.border = border_header()

rules = [
    ("Rule 1: File size", "Компонент <= 150 строк, страница <= 40 строк", "PASS"),
    ("Rule 2: useState limit", "Максимум 3 useState на компонент", "PASS"),
    ("Rule 3: No fetch inside", "Запрещён fetch/axios внутри UI-компонентов", "PASS"),
    ("Rule 4: Barrel exports", "Каждый модуль имеет index.ts barrel export", "PASS"),
    ("Rule 5: Layer separation", "Компоненты импортируют только из tokens/", "PASS"),
    ("forwardRef", "Все примитивы используют forwardRef", "PASS"),
    ("JSDoc + @example", "Каждый компонент документирован с примером", "PASS"),
    ("data-slot", "Каждый компонент имеет data-slot атрибут", "PASS"),
    ("cn() утилита", "Все стили через cn() (clsx + tailwind-merge)", "PASS"),
]

for i, (rule, desc, status) in enumerate(rules, 1):
    row = 5 + i - 1
    ws_rules.cell(row=row, column=2, value=i).font = font_body()
    ws_rules.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws_rules.cell(row=row, column=3, value=rule).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
    ws_rules.cell(row=row, column=4, value=desc).font = font_body()
    sc = ws_rules.cell(row=row, column=5, value=status)
    sc.font = FONT_EXTRACTED
    sc.fill = FILL_EXTRACTED
    sc.alignment = Alignment(horizontal="center", vertical="center")
    fill = fill_data_row(i)
    for col in range(2, 6):
        c = ws_rules.cell(row=row, column=col)
        if col != 5:
            c.fill = fill
        c.border = BORDER_THIN
    ws_rules.row_dimensions[row].height = 24

ws_rules.freeze_panes = "B5"

# ============================================================
# SAVE
# ============================================================
OUTPUT = "/home/z/my-project/download/ui-kit-checklist.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Grand total: {grand_total} components ({grand_extracted + len(SHADCN_PRIMITIVES)} extracted, {grand_planned} planned)")
